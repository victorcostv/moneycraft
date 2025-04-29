from flask import Flask, render_template, request, send_file, jsonify
import fitz  # PyMuPDF
import pandas as pd
import re
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime, timedelta

app = Flask(__name__, template_folder="templates")
UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "outputs"
CATEGORIES_FILE = "categories.json"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Se o arquivo de categorias não existir, criamos um padrão
if not os.path.exists(CATEGORIES_FILE):
    default_categories = {
        "Alimentação": ["Mc Donalds", "Burger King", "KFC", "Restaurante", "Lanchonete", "Pizza", "Subway"],
        "Transporte": ["Uber", "99 Táxi", "Posto", "Gasolina", "Rodoviária", "Metrô"],
        "Compras": ["Mercadolivre", "Shopee", "Magazine Luiza", "Americanas", "Casas Bahia"],
        "Saúde": ["Drogaria", "Raia", "Farmácia", "Consulta", "Hospital", "Clínica"],
        "Entretenimento": ["Netflix", "Spotify", "Cinema", "Ingresso", "Teatro"]
    }
    with open(CATEGORIES_FILE, "w") as f:
        json.dump(default_categories, f, indent=4)

def load_categories():
    with open(CATEGORIES_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_categories(categories):
    with open(CATEGORIES_FILE, "w") as f:
        json.dump(categories, f, indent=4)

@app.route("/")
def home():
    categories = load_categories()
    return render_template("index.html", categories=categories)

@app.route("/upload", methods=["POST"])
def upload_file():
    file = request.files["file"]
    if file:
        pdf_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(pdf_path)
        excel_path = process_pdf(pdf_path)
        return send_file(excel_path, as_attachment=True)
    return "Erro ao processar o arquivo."

@app.route("/categories", methods=["GET"])
def get_categories():
    return jsonify(load_categories())

@app.route("/edit-categories")
def edit_categories():
    categories = load_categories()
    return render_template("categories.html", categories=categories)

@app.route("/update-categories", methods=["POST"])
def update_categories():
    data = request.json
    save_categories(data)
    return jsonify({"message": "Categorias atualizadas com sucesso!"})

def process_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    transactions = []
    current_date = None
    pattern_date = re.compile(r"(\d{2} \w{3})")
    pattern_value = re.compile(r"R\$ -?\d+,\d{2}")
    categories = load_categories()
    
    def classify_expense(description):
        for category, keywords in categories.items():
            if any(keyword.lower() in description.lower() for keyword in keywords):
                return category
        return "Outros"
    
    for page in doc:
        lines = page.get_text("text").split("\n")
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if pattern_date.fullmatch(line):
                current_date = line
            elif pattern_value.fullmatch(line) and current_date:
                value = float(line.replace("R$ ", "").replace(",", "."))
                description = lines[i - 1].strip()
                category = classify_expense(description)
                
                match = re.search(r"(\d+)/(\d+)", description)
                if match:
                    current_installment = int(match.group(1))
                    total_installments = int(match.group(2))
                    clean_description = re.sub(r"-?\s*\d+/\d+", "", description).strip()
                    for future_month in range(total_installments - current_installment + 1):
                        future_date = datetime.now() + timedelta(days=30 * future_month)
                        month_name = future_date.strftime("%b %Y")
                        transactions.append([month_name, current_date, clean_description, value, category, f"{current_installment + future_month}/{total_installments}"])
                else:
                    transactions.append([datetime.now().strftime("%b %Y"), current_date, description, value, category, ""])  
            i += 1
    
    df = pd.DataFrame(transactions, columns=["Mês", "Data", "Descrição", "Valor", "Categoria", "Parcela"])
    output_path = os.path.join(OUTPUT_FOLDER, "fatura_nubank.xlsx")
    
    wb = Workbook()
    sorted_months = sorted(df["Mês"].unique(), key=lambda x: datetime.strptime(x, "%b %Y"))
    
    for month in sorted_months:
        ws = wb.create_sheet(title=month)
        header = ["Data", "Descrição", "Valor", "Categoria", "Parcela"]
        ws.append(header)
        ws.column_dimensions["B"].width = 30  # Define a largura da coluna B para 30
        
        for col in range(1, len(header) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=1, column=col).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
            ws.cell(row=1, column=col).border = Border(bottom=Side(style="thin"))
        
        for row in df[df["Mês"] == month].itertuples(index=False, name=None):
            ws.append(row[1:])
        
        ws.append(["", "Total da Fatura", df[df["Mês"] == month]["Valor"].sum()])
        
    wb.remove(wb["Sheet"])
    wb.save(output_path)
    return output_path

if __name__ == "__main__":
    app.run(debug=True)
